from selenium import webdriver 
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl 

from openpyxl import load_workbook
from openpyxl import Workbook

from openpyxl.utils import get_column_letter, column_index_from_string


# **********************VARIABLES*******************


pathex = "C:\\Users\\myUser\\Desktop\\correlas\\py-xl.xlsx"


url = 'https://www.latlong.net/convert-address-to-lat-long.html'

driver = webdriver.Chrome("C:\\Program Files (x86)\\chromedriver.exe")

driver.get(url)

driver.find_element_by_id('cookiescript_close').click() #closing cookie alert

driver.find_element_by_css_selector('img.icons').click() #enter to My accoun/user name icon



# ****************YOUR CREDENTALS GOES HERE *******************
myemail = 'your-email@whatever.xxx'
mypass = '********'

# *************************************************************

wb = openpyxl.load_workbook('py-xl.xlsx')
sheet = wb['Hoja1']


rows = sheet.min_row
rowM = sheet.max_row

cols = sheet.min_column
colM = sheet.max_column


# **************************EXCEPTIONS*************************

try:
    element = WebDriverWait(driver, 10).until( 
        EC.presence_of_element_located((By.ID, "email")) #Verifying the existence of email box 
    )
    element.click()
    time.sleep(4)

    driver.find_element_by_id('email').send_keys(myemail) 
    time.sleep(4)


    element = WebDriverWait(driver, 10).until( 
        EC.presence_of_element_located((By.ID, "password1")) #Verifying the existence of password box 
    )
    element.click()

    driver.find_element_by_id('password1').send_keys(mypass)
    driver.find_element_by_class_name('button').click()


except:
    print('there was a mistake')
    time.sleep(3)


try:
    element = WebDriverWait(driver, 10).until( 
        EC.presence_of_element_located((By.ID, "myTopnav")) #Verifying the existence of email box 
    )
    element.click()
    driver.find_element_by_link_text('Geographic Tools').click()
    

    element = WebDriverWait(driver, 10).until( 
        EC.presence_of_element_located((By.LINK_TEXT, "Address to Latitude and Longitude")) #Verifying the existence of email box 
    )
    driver.find_element_by_link_text('Address to Latitude and Longitude').click()
    driver.find_element_by_id('cookiescript_close').click() #closing cookie alert
       

except:
    print('Another mistake')
    time.sleep(3)


# ******************* MAIN FOR LOOP ********************************

for r in range(2, rowM+1):
    for c in range(3, 4):       
        y=sheet.cell(r,c)
        x=y.value       

        driver.find_element_by_css_selector('input.width70').send_keys(x)
        time.sleep(2)
        driver.find_element_by_id('btnfind').click()
        time.sleep(5)
        driver.find_element_by_id('lat').click()

        lats = driver.find_element_by_id("lat") # GETTING LATITUDE ****************
        print(lats.get_attribute('value'))
        time.sleep(2)

        longs = driver.find_element_by_id("lng") # GETTING LONGITUDE ****************
        print(longs.get_attribute('value'))

        time.sleep(2)

        lat_main = lats.get_attribute('value')
        long_main= longs.get_attribute('value')
        time.sleep(4)


		# EXPORTING LATITUDE AND LONGITUDE TO EXCEL

        w_b = load_workbook(filename = 'py-xl.xlsx') # where to get & write data

        book = Workbook()

        sheet_ranges = w_b['Hoja1'] # especifying 


        sheet_ranges.cell(row=r,column=9).value = lat_main
        sheet_ranges.cell(row=r,column=10).value = long_main

        w_b.save(filename = 'py-xl.xlsx') # Saving one file in especifically
        driver.find_element_by_css_selector('input.width70').send_keys(Keys.CONTROL, "a")
        # time.sleep(20)
